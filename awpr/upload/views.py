# PR2018-04-14
import json

from django.contrib.auth.decorators import login_required # PR2018-04-01
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect #, get_object_or_404
from django.utils.decorators import method_decorator
#PR2022-02-13 was ugettext_lazy as _, replaced by: gettext_lazy as _
from django.utils.translation import gettext_lazy as _
from django.views.generic import View

from awpr import functions as af
from awpr import settings as s
from awpr import constants as c
from awpr import menus as awpr_menu

from schools import models as sch_mod
from students import models as stud_mod
from subjects import models as subj_mod

import openpyxl
# PR2018-04-27
import logging
logger = logging.getLogger(__name__)

# PR2018-05-06


# PR2018-04-27 import Excel file from "How to upload and process the Excel file in Django" http://thepythondjango.com/upload-process-excel-file-django/



# ========  UPLOAD AWP =====================================
@method_decorator([login_required], name='dispatch')
class UploadOldAwpView(View):  #PR2020-12-13 PR2021-05-03 PR2021-07-03

    def post(self,request):
        logging_on = s.LOGGING_ON

        files = request.FILES
        uploadedfile = files.get('file')

# get instance of examyear from settings
        sel_examyear, selected_dict_has_changed_NIU, multiple_examyears_exist = af.get_sel_examyear_with_default(request)

        if logging_on:
            logger.debug(' ')
            logger.debug(' ============= UploadOldAwpView ============= ')
            logger.debug('file: ' + str(uploadedfile) + ' ' + str(type(uploadedfile)))
            logger.debug('sel_examyear: ' + str(sel_examyear) + ' ' + str(type(sel_examyear)))

        logfile = list()
        # check if request.user.country is parent of sel_examyear_instance PR2018-10-18
        if sel_examyear is not None and request.user.country is not None and \
                sel_examyear.country.id == request.user.country.id:

            wb = openpyxl.load_workbook(uploadedfile)
            # PR2018-04-27 debug: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames). Was:  ws_names = wb.get_sheet_names()

            mapped = {}

            # school instance is retrieved when ws_name = 'school', It keeps value in the rest of the worksheets
            school = None

            # school must be first in list, checks if examyear and country equal selected examyear and country
            ws_list = ('school', 'department', 'level', 'sector', 'subjecttype', 'scheme',
                        'subject', 'schemeitem', 'package', 'packageitem', 'cluster', 'student', 'studsubj')
# -------------------------------------------------------------------------------
# - iterate through ws_list to make sure the data are imported in the right order
            for ws_name in ws_list:

# - lookup worksheet
                index = -1
                for wb_index, wb_sheetname in enumerate(wb.sheetnames):
                    if wb_sheetname == ws_name:
                        index = wb_index
                        break
                worksheet = None
                if index > -1:
                    worksheet = wb.worksheets[index]

                if worksheet:

# ----------------------------------------------------------------------------------
# - iterate over the rows of this worksheet and get the  value from each cell in row

# first row contains column names, put them in list 'column_names'
                    is_first_row = True
                    column_names = list()

                    if logging_on:
                        row_count = 0
                        for row in worksheet.iter_rows():
                            row_count += 1
                        logger.debug('ws_name  : ' + str(ws_name) + ' row_count : ' + str(row_count))

                    row_count = 0
                    for row in worksheet.iter_rows():
                        row_count += 1
                        row_data = {}
                        for i, cell in enumerate(row):
                            if is_first_row:
                                column_names.append(str(cell.value))
                            else:
                                col_name = column_names[i]
# - put cell value in dict 'row_data', with key = col_name
                                if cell.value:
                                    row_data[col_name] = cell.value
                        # column_names: ['scheme_id', 'dep_id', 'dep', 'sct_id', 'sct', 'lvl_id', 'lvl']
                        # row_data: {'scheme_id': 1, 'dep_id': 1, 'dep': 'vsbo', 'sct_id': 1, 'sct': 'tech', 'lvl_id': 1, 'lvl': 'tkl'}

                        if row_data and not is_first_row:
                            if ws_name == 'school':
                                school = ImportSchool(ws_name, row_data, logfile, mapped, sel_examyear, request)
                            elif school:
                                if ws_name == 'department':
                                    ImportDepartment(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'level':
                                    ImportLevel(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'sector':
                                    ImportSector(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'subjecttype':
                                    ImportSubjecttype(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'scheme':
                                    ImportScheme(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'subject':
                                    ImportSubject(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'schemeitem':
                                    ImportSchemeitem(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'package':
                                    pass
                                    # don't import package
                                    #ImportPackage(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'packageitem':
                                    pass
                                    # don't import packageitem
                                    # ImportPackageitem(ws_name, row_data, logfile, mapped, sel_examyear, request)
                                elif ws_name == 'cluster':
                                    pass
                                    # don't import cluster
                                    # ImportCluster(ws_name, row_data, logfile, mapped, sel_examyear, school, request)
                                elif ws_name == 'student':
                                    ImportStudent(ws_name, row_data, logfile, mapped, sel_examyear, school, request)
                                elif ws_name == 'studsubj':
                                    ImportStudentsubject(ws_name, row_data, logfile, mapped, sel_examyear, school, request)

                        is_first_row = False
# - end of loop
# -------------------------------------------------------------------------------

        header_text = _('Upload') + ' ' + str(_('All').lower())
        update_wrap = {"logfile": logfile, 'header': header_text}

# 9. return update_wrap
        return HttpResponse(json.dumps(update_wrap, cls=af.LazyEncoder))

def ImportData(ws_name, row_data, logfile, mapped, sel_examyear, request):  #PR2020-12-13
    logging_on = s.LOGGING_ON

    try:
        requsr_country = request.user.country

        if ws_name == 'birthcountry':
            name = str(row_data[0]) if row_data[0] else None
            # PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
            if name:

                # check if birthcountry already exists
                birthcountry = get_birthcountry(name)
                if birthcountry is None:
                    birthcountry = stud_mod.Birthcountry(
                        name=name
                    )
                    birthcountry.save(request=request)
                    logfile.append(row_data)

        elif ws_name == 'birthplace':
            birthcountry_name = str(row_data[0]) if row_data[0] else None
            birthplace_name = str(row_data[1]) if row_data[1] else None

            birthcountry = get_birthcountry(birthcountry_name)
            if birthcountry and birthplace_name:
                birthplace = get_birthplace(birthcountry, birthplace_name)
                if birthplace is None:
                    birthplace = stud_mod.Birthplace(
                        birthcountry=birthcountry,
                        name=birthplace_name
                    )
                    birthplace.save(request=request)
                    logfile.append(row_data)

        elif ws_name in ('schoolCUR', 'schoolSXM'):
            # 0: country  1: code  2: name  3: abbrev  4: article  5: depbases  6: is_template
            country_code = str(row_data[0]) if row_data[0] else None
            school_code = str(row_data[1]) if row_data[1] else None

            if logging_on:
                logger.debug ('-------------------  school ----------------- examyear: ' + str(sel_examyear))
                logger.debug ('country_code: ' + str(country_code))
                logger.debug ('school_code: ' + str(school_code))

    # - get country based on code 'Cur' in excel file, not requsr_country with this code already exists in this country. If not: create
            exc_country = af.get_country_instance_by_abbrev(country_code)

    # skip if exc_country is different from requsr_country
            if exc_country and requsr_country and exc_country.pk == requsr_country.pk:

    # - check if schoolbase with this code already exists in this country. If not: create
                schoolbase = get_schoolbase(exc_country, school_code)
                if schoolbase is None:
                    schoolbase = sch_mod.Schoolbase.objects.create(country=exc_country, code=school_code)
                    if logging_on:
                        logger.debug('schoolbase created: ' + str(schoolbase))

    # - check if school with this schoolbase already exists in this examyear. If not: create
                school = get_school(schoolbase, sel_examyear)
                if logging_on:
                    logger.debug('school found: ' + str(school))

                if schoolbase is not None and school is None:
                    # 0: country  1: code  2: name  3: abbrev  4: article  5: depbases  6: is_template
                    name = str(row_data[2]) if row_data[2] else None
                    abbrev = str(row_data[3]) if row_data[3] else None
                    article = str(row_data[4]) if row_data[4] else None
                    depbases = get_depbase_id_list_from_nameslist(row_data[5], mapped)
                    # is_template = True if str(row_data[6]) == "1" else False

                    # PR2018-04-28 debug: don't forget de brackets when creating an instance of the class
                    school = sch_mod.School(
                        base=schoolbase,
                        examyear=sel_examyear,
                        name=name,
                        abbrev=abbrev,
                        article=article,
                        depbases=depbases
                    )
                    if logging_on:
                        logger.debug('school: ' + str(school))
                    school.save(request=request)
                logfile.append(row_data)
    except Exception as e:
        logger.error(getattr(e, 'message', str(e)))
        row_data[0] = _("An error occurred. '%(fld)s' is not saved.") % {'fld': ws_name}
        logfile.append(row_data)
# - end of ImportData
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@


def ImportDepartment(ws_name, row_data, logfile, mapped, sel_examyear, request):  # PR2021-05-03 PR021-07-11

    logging_on = False  # s.LOGGING_ON
    if ws_name == 'department' and row_data:
        try:
            if logging_on:
                logger.debug ('-------------------  department ----------------- sel_examyear: ' + str(sel_examyear))
                logger.debug ('row_data: ' + str(row_data))
            # row_data: {'department_id': 1, 'code': 'vsbo', 'abbrev': 'V.S.B.O.', 'name': 'Voorbereidend Secundair Beroepsonderwijs'}

            code = row_data.get('code')
            if code:

    # - check if depbase with this code already exists .
                depbase = sch_mod.Departmentbase.objects.filter(
                    code__iexact=code
                ).order_by('pk').first()
            # - create depbase if it does not exist
                if depbase is None:
                    # 'vsbo' becomes 'Vsbo'
                    code_capitalized = code.capitalize()
                    depbase = sch_mod.Departmentbase(
                        code=code_capitalized)
                    depbase.save()
                    logfile.append('depbase created: ' + str(depbase))
                    if logging_on:
                        logger.debug ('depbase created: ' + str(depbase))

    # - check if department with this depbase already exists in this examyear. If not: create
                department = sch_mod.Department.objects.filter(
                    base=depbase,
                    examyear=sel_examyear
                ).order_by('pk').first()

                if department is None:
                    name = row_data.get('name')
                    abbrev = row_data.get('abbrev')
                    sequence = 1 if (code == 'vsbo') else 2 if (code == 'havo') else 3 if (code == 'vwo') else 4
                    level_req = (code == 'vsbo')
                    sector_req = True
                    has_profiel = (code in ('havo', 'vwo'))

                    department = sch_mod.Department(
                        base=depbase,
                        examyear=sel_examyear,
                        name=name,
                        abbrev=abbrev,
                        sequence=sequence,
                        level_req=level_req,
                        sector_req=sector_req,
                        has_profiel=has_profiel
                    )
                    department.save(request=request)
                    logfile.append('depbase created: ' + str(department))

                    if logging_on:
                        logger.debug ('depbase created = ' + str(department))

                if department:
                    awp_department_id = row_data.get('department_id')
                    if 'department' not in mapped:
                        mapped['department'] = {}
                    mapped['department'][awp_department_id] = department.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error department: ' + str(e))


def ImportLevel(ws_name, row_data, logfile, mapped, sel_examyear, request):  #PR2021-05-03

    logging_on = False  # s.LOGGING_ON
    if ws_name == 'level' and row_data:
        try:
            # AWP table 'Studierichting' also contains profielen. Is filtered out in AWP
            if logging_on:
                logger.debug('-------------------  level ----------------- sel_examyear: ' + str(sel_examyear))
                logger.debug('row_data: ' + str(row_data))
                #  row_data: {'level_id': 1, 'abbrev': 'TKL', 'name': 'Theoretisch Kadergerichte Leerweg', 'depbases': '1;'}

            abbrev = row_data.get('abbrev')
            depbases = row_data.get('depbases')

            if abbrev and depbases:
    # - check if lvlbase with this code (abbrev) already exists
                lvlbase = subj_mod.Levelbase.objects.filter(
                    code__iexact=abbrev
                ).order_by('pk').first()

        # - create lvlbase if it does not exist
                if lvlbase is None:
                    # 'pkl' becomes 'PKL'
                    code_uc = abbrev.upper()
                    lvlbase = subj_mod.Levelbase(
                        code=code_uc
                    )
                    lvlbase.save()
                    logfile.append('lvlbase created: ' + str(lvlbase))
                    if logging_on:
                        logger.debug('lvlbase created: ' + str(lvlbase))

    # - check if level with this lvlbase already exists in this examyear.
                level = subj_mod.Level.objects.filter(
                    base=lvlbase,
                    examyear=sel_examyear
                ).order_by('pk').first()

    # - if levelbase does not exist: create new level record
                if level is None:
                    name = row_data.get('name')
                    sequence = 1 if (abbrev == 'PBL') else 2 if (abbrev == 'PKL') else 3 if (abbrev == 'TKL') else 4

                    level = subj_mod.Level(
                        base=lvlbase,
                        examyear=sel_examyear,
                        name=name,
                        abbrev=lvlbase.code,
                        sequence=sequence,
                        depbases=depbases
                    )
                    level.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(level))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(level))

                if level:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(level))
                    awp_level_id = row_data.get('level_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_level_id] = level.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error level: ' + str(e))
# - end of ImportLevel


def ImportSector(ws_name, row_data, logfile, mapped, sel_examyear, request):  #PR2021-05-03

    logging_on = False  # s.LOGGING_ON
    if ws_name == 'sector' and row_data:
        try:
            if logging_on:
                logger.debug('-------------------  sector ----------------- sel_examyear: ' + str(sel_examyear))
                logger.debug('row_data: ' + str(row_data))
                # row_data: {'level_id': 1, 'abbrev': 'TKL', 'name': 'Theoretisch Kadergerichte Leerweg', 'depbases': '1;'}

            abbrev = row_data.get('abbrev')
            depbases = row_data.get('depbases')

            if abbrev and depbases:
    # AWP table 'Studierichting' also contains profielen. Is filtered out in AWP

    # - check if sctbase with this code (abbrev) already exists
                sctbase = subj_mod.Sectorbase.objects.filter(
                    code__iexact=abbrev
                ).order_by('pk').first()

    # - create sctbase if it does not exist
                if sctbase is None:
                    sctbase = subj_mod.Sectorbase(
                        code=abbrev
                    )
                    sctbase.save()
                    logfile.append('sctbase created: ' + str(sctbase))
                    if logging_on:
                        logger.debug('sctbase created: ' + str(sctbase))

    # - check if sector with this sctbase already exists in this examyear.
                sector = subj_mod.Sector.objects.filter(
                    base=sctbase,
                    examyear=sel_examyear
                ).order_by('pk').first()

    # - create new sector record
                if sector is None:
                    name = row_data.get('name')
                    sequence = 1 if (abbrev == 'PBL') else 2 if (abbrev == 'PKL') else 3 if (abbrev == 'TKL') else 4

                    sector = subj_mod.Sector(
                        base=sctbase,
                        examyear=sel_examyear,
                        name=name,
                        abbrev=abbrev,
                        sequence=sequence,
                        depbases=depbases
                    )
                    sector.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(sector))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(sector))

                if sector:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(sector))
                    awp_sector_id = row_data.get('sector_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_sector_id] = sector.pk

                if logging_on:
                    logger.debug ('mapped: ' + str(mapped))

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error sector: ' + str(e))
# - end of ImportSector


def ImportSubjecttype(ws_name, row_data, logfile, mapped, sel_examyear, request):  #PR2021-05-03

    logging_on = False  # s.LOGGING_ON
    if ws_name == 'subjecttype' and row_data:
        try:
            if logging_on:
                logger.debug('-------------------  subjecttype ----------------- sel_examyear: ' + str(sel_examyear))
                logger.debug('row_data: ' + str(row_data))
            #  row_data: {'subjecttype_id': 1, 'department_id': 1, 'name': 'Gemeenschappelijk deel'}

            awp_name = row_data.get('name')
            if awp_name:
                awp_name_sliced_lc = awp_name.lower()[0:8]

                code_dict = {'gemeensc': 'gmd', 'sectorde': 'spd', 'profield': 'spd', 'overig v': 'vrd',
                             'vrije de': 'vrd', 'sectorpr': 'spr', 'sectorwe': 'wst', 'profielw': 'wst', 'stage': 'stg'}
                name_dict = {'gemeensc': 'Gemeenschappelijk deel', 'sectorde': 'Sectordeel', 'profield': 'Profieldeel',
                             'overig v': 'Overig vak', 'vrije de': 'Vrije deel', 'sectorpr': 'Sectorprogramma',
                             'sectorwe': 'Sectorwerkstuk', 'profielw': 'Profielwerkstuk', 'stage': 'Stage'}
                abbrev_dict = {'gemeensc': 'Gemeensch.', 'sectorde': 'Sectordeel', 'profield': 'Profieldeel',
                               'overig v': 'Overig vak', 'vrije de': 'Vrije deel', 'sectorpr': 'Sectorprog.',
                               'sectorwe': 'Werkstuk', 'profielw': 'Werkstuk', 'stage': 'Stage'}
                sequence_dict = {'gemeensc': 1, 'sectorde': 2, 'profield': 2, 'overig v': 3, 'vrije de': 3,
                                 'sectorpr': 4, 'sectorwe': 5, 'profielw': 5, 'stage': 6}

                code = code_dict.get(awp_name_sliced_lc)
                name = name_dict.get(awp_name_sliced_lc)
                abbrev = abbrev_dict.get(awp_name_sliced_lc)
                sequence = sequence_dict.get(awp_name_sliced_lc, 1)

                if logging_on:
                    logger.debug('awp_name: ' + str(awp_name))
                    logger.debug('awp_name_sliced_lc: ' + str(awp_name_sliced_lc))
                    logger.debug('code: ' + str(code))
                    logger.debug('name: ' + str(name))
                    logger.debug('abbrev: ' + str(abbrev))
                    logger.debug('sequence: ' + str(sequence))

                sjtpbase = None
                if code:
    # - check if subjecttypebase with this code already exists
                    sjtpbase = subj_mod.Subjecttypebase.objects.filter(
                        code__iexact=code
                    ).order_by('pk').first()

                    if logging_on:
                        logger.debug('existing sjtpbase: ' + str(sjtpbase))

        # - create subjecttypebase if it does not exist
                    if sjtpbase is None:
                        if name and abbrev:
                            sjtpbase = subj_mod.Subjecttypebase.objects.create(
                                code=code,
                                name=name,
                                abbrev=abbrev,
                                sequence=sequence
                            )
                            if logging_on:
                                logger.debug('created sjtpbase: ' + str(sjtpbase))

# - create subjecttype for each scheme, if it doesn't already exist
                if sjtpbase:
                    schemes = subj_mod.Scheme.objects.filter(
                        department__examyear=sel_examyear
                    )
                    for scheme in schemes:
                        if logging_on:
                            logger.debug('scheme: ' + str(scheme))

                        subjecttype = subj_mod.Subjecttype.objects.filter(
                            base=sjtpbase,
                            scheme=scheme
                        ).order_by('pk').first()

                        if logging_on:
                            logger.debug('existing subjecttype: ' + str(subjecttype))

                        if subjecttype is None:
                            subjecttype = create_upload_subjecttype(sjtpbase, scheme)
                            if logging_on:
                                logger.debug ('created subjecttype: ' + str(subjecttype))

                    if sjtpbase:
                        awp_id = row_data.get('subjecttype_id', '')
                        if 'sjtpbase' not in mapped:
                            mapped['sjtpbase'] = {}
                        mapped['sjtpbase'][awp_id] = sjtpbase.pk
                        if logging_on:
                            logger.debug('mapped sjtpbase: ' + str( mapped['sjtpbase']))
        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error subjecttype: ' + str(e))

# - end of ImportSubjecttype


def create_upload_subjecttype(sjtpbase, scheme):
    # create a new subjecttype when uploading from awp PR2021-07-03

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('------  create_upload_subjecttype ----------')
        logger.debug('   sjtpbase: ' + str(sjtpbase))
        logger.debug('   scheme: ' + str(scheme))

    subjecttype = None

    try:
        depbase_code = scheme.department.base.code
        if logging_on:
            logger.debug('depbase_code: ' + str(depbase_code))

        lvl_abbrev = None
        if scheme.level:
            lvl_abbrev = scheme.level.abbrev

        if logging_on:
            logger.debug('lvl_abbrev: ' + str(lvl_abbrev))
            logger.debug('sjtpbase.code: ' + str(sjtpbase.code))
        # stage only in Vsbo - pbl pkl
        # werkstuk only in and Havo, Vwo and Vsbo tkl
        # sectorprog only in Vsbo


        skip = False
        if sjtpbase.code == 'stg':
            skip = lvl_abbrev is None or lvl_abbrev == 'TKL'
        elif sjtpbase.code == 'wst':
            skip = lvl_abbrev and lvl_abbrev in ('PBL', 'PKL')
        elif sjtpbase.code == 'spr':
            skip = depbase_code in ('Havo', 'Vwo')
        if not skip:
            name = None
            abbrev = None
            has_prac = False
            has_pws = False
            if sjtpbase.code == 'gmd':
                name = 'Gemeenschappelijk deel'
                abbrev = 'Gemeensch.'
            if sjtpbase.code == 'spd':
                name = 'Sectordeel' if depbase_code == 'Vsbo' else 'Profieldeel'
                abbrev = name
            elif sjtpbase.code == 'vrd':
                name = 'Overig vak' if depbase_code == 'Vsbo' else 'Vrije deel'
                abbrev = name
            elif sjtpbase.code == 'spr':
                name = 'Sectorprogramma'
                abbrev = 'Sectorprog.'
                has_prac = True
            elif sjtpbase.code == 'wst':
                name = 'Sectorwerkstuk' if depbase_code == 'Vsbo' else 'Profielwerkstuk'
                abbrev = 'Werkstuk'
                has_pws = True
            elif sjtpbase.code == 'stg':
                name = 'Stage'
                abbrev = name

            if logging_on:
                logger.debug('name: ' + str(name))
                logger.debug('abbrev: ' + str(abbrev))
                logger.debug('has_prac: ' + str(has_prac))
                logger.debug('has_pws: ' + str(has_pws))
            # create function also saves the new instance
            subjecttype = subj_mod.Subjecttype.objects.create(
                base=sjtpbase,
                scheme=scheme,
                name=name,
                abbrev=abbrev,
                has_prac=has_prac,
                has_pws=has_pws
            )

    except Exception as e:
        logger.error(getattr(e, 'message', str(e)))

    return subjecttype


def ImportScheme(ws_name, row_data, logfile, mapped, examyear, request):  #PR2021-05-04

    logging_on = False  # s.LOGGING_ON

    # row_data: {'scheme_id': 9, 'department_id': 1, 'sector_id': 3, 'level_id': 3}
    if ws_name == 'scheme' and row_data:
        try:
            department = get_department_from_mapped(row_data, examyear, mapped)
            level = get_level_from_mapped(row_data, examyear, department, mapped)
            sector = get_sector_from_mapped(row_data, examyear, department, mapped)

            if logging_on:
                logger.debug('-------------------  scheme ----------------- sel_examyear: ' + str(examyear))
                logger.debug('row_data: ' + str(row_data))
                logger.debug('department: ' + str(department))
                logger.debug('level: ' + str(level))
                logger.debug('sector: ' + str(sector))

            if department and sector:

    # - check if scheme already exists in this examyear. If not: create
                scheme = subj_mod.Scheme.objects.filter(
                    department=department,
                    level=level,
                    sector=sector
                ).order_by('pk').first()

    # - create new scheme record
                if scheme is None:
                    name_arr = []
                    if department.base:
                        name_arr.append(department.base.code)
                    if level:
                        name_arr.append(level.abbrev)
                    if sector:
                        name_arr.append(sector.abbrev)
                    name = ' - '.join(name_arr)

                    fields_arr = ['mand', 'comb']
                    dep_code = department.base.code.lower()
                    if dep_code == 'vwo':
                        fields_arr.append('chal')
                    elif dep_code == 'vsbo':
                        if level:
                            lvl_abbrev = level.abbrev.lower()
                            if lvl_abbrev in ('pbl', 'pkl'):
                                fields_arr.append('prac')
                    fields = ';'.join(fields_arr)

                    scheme = subj_mod.Scheme(
                        department=department,
                        level=level,
                        sector=sector,
                        name=name,
                        fields=fields
                    )
                    scheme.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(scheme))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(scheme))

                if scheme:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(scheme))
                    awp_scheme_id = row_data.get('scheme_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_scheme_id] = scheme.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error scheme: ' + str(e))
# - end of ImportScheme


def ImportSubject(ws_name, row_data, logfile, mapped, examyear, request):  #PR2021-05-04

    logging_on = s.LOGGING_ON
    if ws_name == 'subject' and row_data:
        try:
            # row_data keys: subject_id, code, name, sequence, depbases, addedbyuser
            code = row_data.get('code')
            name = row_data.get('name')

            if code:
                if logging_on:
                    logger.debug('-------------------  subject ----------------- ')
                    logger.debug('row_data: ' + str(row_data))
                    logger.debug('code: ' + str(code))
                    logger.debug('name: ' + str(name))
                    addedbyuser = row_data.get('addedbyuser')
                    logger.debug('addedbyuser: ' + str(addedbyuser) + ' ' + str(type(addedbyuser)))
                    #  row_data: {'level_id': 1, 'abbrev': 'TKL', 'name': 'Theoretisch Kadergerichte Leerweg', 'depbases': '1;'}

        # - check if subject with this code already exists in this examyear. If not: create
                # subjectbase codes have changed in AWP-online. Map old code to new ones, to link with correct subjectbase
            # - first check if subject with new_subj_code exists
                new_subj_code = get_mapped_subject_code(code, name)
                subject = subj_mod.Subject.objects.filter(
                    examyear=examyear,
                    base__code__iexact=new_subj_code
                ).order_by('pk').first()
            # - if not found: check if subject with old code existst
                if subject is None:
                    subject = subj_mod.Subject.objects.filter(
                        examyear=examyear,
                        base__code__iexact=code
                    ).order_by('pk').first()
            # - if not found: create new subject record with new code
                if subject is None:

    # - first create new base record with new code
                    base = subj_mod.Subjectbase(
                        code=new_subj_code
                    )
                    base.save()

                    name = row_data.get('name')
                    sequence = row_data.get('sequence')
                    if sequence is None:
                        sequence = 9999
                    # NIU addedbyschool = True if row_data.get('addedbyuser') == 1 else False
                    depbases = row_data.get('depbases')

                    subject = subj_mod.Subject(
                        base=base,
                        examyear=examyear,
                        name=name,
                        sequence=sequence,
                        depbases=depbases,
                        addedbyschool=False
                    )
                    subject.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(subject))

                    if logging_on:
                        logger.debug(ws_name + ' created = ' + str(subject))

                if subject:
                    if logging_on:
                        logger.debug(ws_name + ': ' + str(subject))
                    awp_subject_id = row_data.get('subject_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_subject_id] = subject.pk

                if logging_on:
                    logger.debug('mapped: ' + str(mapped))

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error subject: ' + str(e))

# - end of ImportSubject


def get_mapped_subject_code(old_awp_code, old_awp_name):  # PR2021-09-03
    # subjectbase codes have changed in AWP-online. Map old code to new ones, to link with correct subjectbase

    if old_awp_code == 'adm&co':
        new_subj_code = 'ac'
    elif old_awp_code == 'amv':
        new_subj_code = 'am'
    elif old_awp_code == 'inst':
        new_subj_code = 'ie'
    elif old_awp_code == 'tech alg':
        new_subj_code = 'ta'
    elif old_awp_code == 'ns1':
        new_subj_code = 'nask1'
    elif old_awp_code == 'ns2':
        new_subj_code = 'nask2'
    elif old_awp_code == 'be':
        new_subj_code = 'bec'
    elif old_awp_code == 'en':
        new_subj_code = 'entl' if 'lit' in old_awp_name else 'en'
    elif old_awp_code == 'fr':
        new_subj_code = 'frtl' if 'lit' in old_awp_name else 'fr'
    elif old_awp_code == 'ne':
        new_subj_code = 'netl' if 'lit' in old_awp_name else 'ne'
    elif old_awp_code == 'sp':
        new_subj_code = 'sptl' if 'lit' in old_awp_name else 'sp'
    else:
        new_subj_code = old_awp_code

    return new_subj_code


def ImportSchemeitem(ws_name, row_data, logfile, mapped, examyear_instance, request):  #PR2021-05-04

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('-------------------  schemeitem ----------------- sel_examyear: ' + str(examyear_instance))
        logger.debug('row_data: ' + str(row_data))
        logger.debug('ws_name: ' + str(ws_name))
        logger.debug('row_data.get(is_mandatory): ' + str(row_data.get('is_mandatory')) + ' ' + str(type(row_data.get('is_mandatory'))))
    # row_data: {'schemeitem_id': 1, 'scheme_id': 1, 'subject_id': 2, 'subjecttype_id': 1, 'gradetype_id': 1,
    #  'weight_se': 1, 'weight_ce': 1, 'is_mandatory': 1  is_combi	has_practexam}

    # fields of schemitem are:
    # scheme, subject, subjecttype, norm ,
    # gradetype, weight_se, weight_ce,
    # is_mandatory,  is_combi, extra_count_allowed,  extra_nocount_allowed, has_practexam,
    # sr_allowed,  reex_combi_allowed, no_centralexam, no_reex, no_thirdperiod, no_exemption_ce,

    if ws_name == 'schemeitem' and row_data:
        try:
            scheme = get_scheme_from_mapped(row_data, examyear_instance, mapped)
            if logging_on:
                logger.debug('scheme: ' + str(scheme))

            subject = get_subject_from_mapped(row_data, examyear_instance, mapped)
            if logging_on:
                logger.debug('subject: ' + str(subject))

            subjecttype = get_subjecttype_from_mapped(row_data, scheme, mapped)
            if logging_on:
                logger.debug('subjecttype: ' + str(subjecttype))

            if scheme and subject and subjecttype:

    # - check if schemeitem already exists in this examyear_instance. If not: create
                schemeitem = subj_mod.Schemeitem.objects.filter(
                    scheme=scheme,
                    subject=subject,
                    subjecttype=subjecttype
                ).order_by('pk').first()

    # - create new schemeitem record
                if schemeitem is None:
                    gradetype = row_data.get('gradetype_id')
                    if gradetype is None:
                        gradetype = 0
                    weight_se = row_data.get('weight_se')
                    if weight_se is None:
                        weight_se = 0
                    weight_ce = row_data.get('weight_ce')
                    if weight_ce is None:
                        weight_ce = 0

                    is_mandatory = True if row_data.get('is_mandatory') == 1 else False
                    #is_manda_subj = True if row_data.get('is_manda_subj') == 1 else False
                    is_combi = True if row_data.get('is_combi') == 1 else False
                    has_practexam = True if row_data.get('has_practexam') == 1 else False

                    schemeitem = subj_mod.Schemeitem(
                        scheme=scheme,
                        subject=subject,
                        subjecttype=subjecttype,
                        gradetype=gradetype,
                        weight_se=weight_se,
                        weight_ce=weight_ce,
                        is_mandatory=is_mandatory,
                        #is_mand_subj=is_mand_subj,
                        is_combi=is_combi,
                        has_practexam=has_practexam
                    )
                    schemeitem.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(schemeitem))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(schemeitem))

                if schemeitem:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(schemeitem))
                    awp_schemeitem_id = row_data.get('schemeitem_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_schemeitem_id] = schemeitem.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error schemeitem: ' + str(e))

        if logging_on:
            logger.debug('mapped: ' + str(mapped))
# - end of Importschemeitem


def ImportPackage(ws_name, row_data, logfile, mapped, examyear, request):  #PR2021-05-05

    logging_on = False  # s.LOGGING_ON

    # row_data: {'scheme_id': 9, 'department_id': 1, 'sector_id': 3, 'level_id': 3}
    if ws_name == 'package' and row_data:
        try:
            scheme = get_scheme_from_mapped(row_data, examyear, mapped)
            name = row_data.get('name')

            if logging_on:
                logger.debug('-------------------  package ----------------- sel_examyear: ' + str(examyear))
                logger.debug('row_data: ' + str(row_data))
                logger.debug('scheme: ' + str(scheme))
                logger.debug('name: ' + str(name))

            if scheme and name:

    # - check if package already exists in this examyear. If not: create
                package = subj_mod.Package.objects.filter(
                    scheme=scheme,
                    name__iexact=name
                ).order_by('pk').first()

    # - create new package record
                if package is None:
                    package = subj_mod.Package(
                        scheme=scheme,
                        name=name
                    )
                    package.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(package))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(package))

                if package:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(package))
                    awp_package_id = row_data.get('package_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_package_id] = package.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error package: ' + str(e))
# - end of ImportPackage


def ImportPackageitem(ws_name, row_data, logfile, mapped, examyear, request):  #PR2021-05-04

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('-------------------  packageitem ----------------- sel_examyear: ' + str(examyear))
        logger.debug('row_data: ' + str(row_data))
        logger.debug('ws_name: ' + str(ws_name))
        # row_data: {'package_id': 1, 'subject_id': 2, 'schemeitem_id': 134}

    if ws_name == 'packageitem' and row_data:
        try:
            package = get_package_from_mapped(row_data, mapped)
            if logging_on:
                logger.debug('package: ' + str(package))
            schemeitem = get_schemeitem_from_mapped(row_data, examyear, mapped)
            if logging_on:
                logger.debug('schemeitem: ' + str(schemeitem))

            if package and schemeitem:
    # - check if packageitem already exists in this examyear. If not: create
                packageitem = subj_mod.Packageitem.objects.filter(
                    package=package,
                    schemeitem=schemeitem
                ).order_by('pk').first()

    # - create new packageitem record
                if packageitem is None:

                    packageitem = subj_mod.Packageitem(
                        package=package,
                        schemeitem = schemeitem
                    )
                    packageitem.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(packageitem))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(packageitem))

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error packageitem: ' + str(e))
# - end of Importpackageitem


def ImportSchool(ws_name, row_data, logfile, mapped, examyear, request):  #PR2021-05-05

    logging_on = s.LOGGING_ON

    school = None
    if ws_name == 'school' and row_data:
        try:
            requsr_country = request.user.country
            # row_data: {'country': 'c', 'exyr': 2021, 'id': 1, 'code': 'cur17', 'name': 'radulphus
            country = row_data.get('country')
            exyr = row_data.get('exyr')
            code = row_data.get('code')

            if code and exyr:
                if logging_on:
                    logger.debug('-------------------  school ----------------- ')
                    logger.debug('row_data: ' + str(row_data))
                    logger.debug('country: ' + str(country))
                    logger.debug('requsr_country.name[0:1].lower(): ' + str(requsr_country.name[0:1].lower()))
                    logger.debug('code: ' + str(code))
                    logger.debug('exyr: ' + str(exyr) + ' ' + str(type(exyr)))
                    logger.debug('examyear.code: ' + str(examyear.code) + ' ' + str(type(examyear.code)))
                    #  row_data: {'country	exyr	id	code	name	depbases}

        # - check if country of this school equals selected country.
                if requsr_country.name[0:1].lower() != country.lower():
                    if logging_on:
                        logger.debug('AWP country: ' + str(c) + ' different from selected country: ' + str(requsr_country.name))

        # - check if examyear of this school equals selected examyear.
                # TODO this is left out for testing
                # if exyr != examyear.code:
                #    if logging_on:
                #        logger.debug('AWP examyear: ' + str(exyr) + ' different fro selected examyear: ' + str(examyear.code))
                # else:
                else:

        # - check if schoolbase with this code already exists. If not: create
                    schoolbase = sch_mod.Schoolbase.objects.filter(
                        country=requsr_country,
                        code__iexact=code
                    ).order_by('pk').first()
                    if logging_on:
                        logger.debug('schoolbase exists: ' + str(schoolbase))
        # - create new schoolbase record
                    if schoolbase is None:
            # - first create new schoolbase record.
                        schoolbase = sch_mod.Schoolbase(
                            country=requsr_country,
                            code=code
                        )

                    if schoolbase:
             # - check if school with this schoolbase already exists in this examyear. If not: create
                        school = sch_mod.School.objects.filter(
                            base=schoolbase,
                            examyear=examyear
                        ).order_by('pk').first()
                        if logging_on:
                            logger.debug('school exists: ' + str(school))

            # - create new school record
                        if school is None:
                # - create new school record.
                            name = row_data.get('name')
                            abbrev = name[0:c.MAX_LENGTH_SCHOOLABBREV]

                            depbases = None
                            awp_depbases = row_data.get('depbases')
                            if awp_depbases:
                                awp_depbases_list = awp_depbases.split(';')
                                if awp_depbases_list:
                                    depbases_list = []

                                    for depbase_str in awp_depbases_list:
                                        if depbase_str:
                                            depbases_list.append(depbase_str)
                                    if depbases_list:
                                        depbases = ';'.join(depbases_list)

                            school = sch_mod.School(
                                base=schoolbase,
                                examyear=examyear,
                                name=name,
                                abbrev=abbrev,
                                # article=article,
                                depbases=depbases
                                #isdayschool
                                #iseveningschool
                                #islexschool
                                #activated
                                #locked
                                #activatedat
                                #lockedat
                            )
                            school.save(request=request)
                            logfile.append(ws_name + ' created: ' + str(school))

                            if logging_on:
                                logger.debug(ws_name + ' created school = ' + str(school))

                        if school:
                            if logging_on:
                                logger.debug(ws_name + ': ' + str(school))
                            awp_school_id = row_data.get('school_id')

                            if ws_name not in mapped:
                                mapped[ws_name] = {}
                            mapped[ws_name][awp_school_id] = school.pk
                            if logging_on:
                                logger.debug('mapped[ws_name]: ' + str(mapped[ws_name]))

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error school: ' + str(e))

    return school
# - end of Importschool


def ImportCluster(ws_name, row_data, logfile, mapped, examyear, school, request):  #PR2021-05-19

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('-------------------  cluster ----------------- sel_examyear: ' + str(examyear))
        logger.debug('row_data: ' + str(row_data))
        logger.debug('ws_name: ' + str(ws_name))
        # row_data: {'cluster_name': 1, 'subject_id': 64, 'dep_id': 3, 'name': 'ne - 1'}

    if ws_name == 'cluster' and row_data:
        try:
            subject = get_subject_from_mapped(row_data, examyear, mapped)
            department = get_department_from_mapped(row_data, examyear, mapped)

            if logging_on:
                logger.debug('subject: ' + str(subject))
                logger.debug('department: ' + str(department))

            if subject and department:
                name = row_data.get('name')

    # - check if cluster already exists in this examyear. If not: create
                cluster = subj_mod.Cluster.objects.filter(
                    school=school,
                    department=department,
                    subject=subject,
                    name__iexact=name
                ).order_by('pk').first()

    # - create new cluster record
                if cluster is None:
                    cluster = subj_mod.Cluster(
                        school=school,
                        department=department,
                        subject=subject,
                        name=name
                    )
                    cluster.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(cluster))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(cluster))

                if cluster:
                    if logging_on:
                        logger.debug (ws_name + ': ' + str(cluster))
                    awp_cluster_id = row_data.get('cluster_id')

                    if ws_name not in mapped:
                        mapped[ws_name] = {}
                    mapped[ws_name][awp_cluster_id] = cluster.pk

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error cluster: ' + str(e))
# - end of ImportCluster


def ImportStudent(ws_name, row_data, logfile, mapped, examyear, school, request):  #PR2021-05-19

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('-------------------  student ----------------- sel_examyear: ' + str(examyear))
        logger.debug('row_data: ' + str(row_data))
        logger.debug('ws_name: ' + str(ws_name))
    if ws_name == 'student' and row_data:
        try:
            department = get_department_from_mapped(row_data, examyear, mapped)

            examyear_int = row_data.get('examyear')
            examyear_instance = None
            if examyear_int == department.examyear.code:
                examyear_instance = department.examyear

            if logging_on:
                logger.debug('school: ' + str(school) + ' ' + str(type(school)))
                logger.debug('department: ' + str(department) + ' ' + str(type(department)))
                logger.debug('examyear_instance: ' + str(examyear_instance) + ' ' + str(type(examyear_instance)))

            if examyear_instance and school and department:
                level = get_level_from_mapped(row_data, examyear_instance, department, mapped)
                sector = get_sector_from_mapped(row_data, examyear_instance, department, mapped)
                scheme = get_scheme_from_mapped(row_data, examyear_instance, mapped)
                package = get_package_from_mapped(row_data, mapped)

                if logging_on:
                    logger.debug('scheme: ' + str(scheme) + ' ' + str(type(scheme)))
                    logger.debug('package: ' + str(package) + ' ' + str(type(package)))

                idnumber = row_data.get('idnumber')

                if idnumber is None:
                    msg_str = 'idnumber is blank'
                    logfile.append(msg_str)
                    if logging_on:
                        logger.debug(msg_str)
                else:
    # - check if student already exists in this school and department
                    student = stud_mod.Student.objects.filter(
                        school=school,
                        department=department,
                        idnumber__iexact=idnumber
                    ).order_by('pk').first()

        # check if department, level and sector of scheme are the same as those of the student
                    is_ok = False
                    if scheme and sector:  #  Havo/Vwo don't have level
                        if department.pk == scheme.department.pk:
                            if sector.pk == scheme.sector.pk:
                                if level is None:
                                    is_ok = (scheme.level is None)
                                else:
                                    is_ok = (level.pk == scheme.level.pk)
                    if not is_ok:
                        msg_str = 'department: ' + str(department) + ', level: ' + str(level) + \
                                       ' and sector: ' + str(sector) + \
                                       ' are not corresponding with scheme: ' + str(scheme)
                        logfile.append(msg_str)
                        if logging_on:
                            logger.debug(msg_str)
                    else:
        # - create new student record if not already exists
                        if student is None:
                    # - first create and save new base record.
                            base = stud_mod.Studentbase.objects.create()

                    # - create new student record
                            student = stud_mod.Student(
                                base=base,
                                school=school,
                                department=department,
                                idnumber=idnumber
                            )
                            student.save(request=request)
                            logfile.append(ws_name + ' created: ' + str(student))
                            if logging_on:
                                logger.debug('student: ' + str(student) + ' ' + str(type(student)))

        # - add info to student instance, update if it is an existing student
                        if student:
                            student.level = level
                            student.sector = sector
                            student.scheme = scheme
                            student.package = package

                            student.lastname = row_data.get('lastname')
                            student.firstname = row_data.get('firstname')
                            student.prefix = row_data.get('prefix')
                            student.gender = row_data.get('gender').upper() if row_data.get('gender') else None

                            student.birthdate = af.get_date_from_ISO(row_data.get('dob'))
                            student.birthcountry = row_data.get('cob')
                            student.birthcity = row_data.get('pob')

                            student.classname = row_data.get('class')
                            student.examnumber = row_data.get('exnr')
                            student.regnumber = row_data.get('regnr')
                            student.diplomanumber = row_data.get('dipnr')
                            student.gradelistnumber = row_data.get('gradelistnr')

                            #student.extrafacilities
                            student.iseveningstudent = school.iseveningschool
                            student.islexstudent = school.islexschool
                            student.bis_exam = True if row_data.get('bis_exam') == 1 else False

                            student.reex_count = row_data.get('reex_count', 0)
                            student.reex03_count = row_data.get('reex03_count', 0)
                            student.has_sere = True if row_data.get('has_sere') == 1 else False
                            student.withdrawn = True if row_data.get('withdrawn') == 1 else False

                            student.gl_ce_avg = row_data.get('gl_ce_avg')
                            student.gl_combi_avg = row_data.get('gl_combi_avg')
                            student.gl_final_avg = row_data.get('gl_final_avg')

                            student.result_info = row_data.get('result_info', 0)

                            student.save(request=request)

                    if student:
                        awp_student_id = row_data.get('student_id')

                        pws_title = row_data.get('pws_title')
                        pws_subjects = row_data.get('pws_subjects')

                        if logging_on:
                            logger.debug (pws_title + ': ' + str(pws_title) + ' ' + str(type(pws_title)))
                            logger.debug (pws_subjects + ': ' + str(pws_subjects) + ' ' + str(type(pws_subjects)))

                        if ws_name not in mapped:
                            mapped[ws_name] = {}

                        mapped[ws_name][awp_student_id] = {
                            'pk': student.pk,
                            'pws_title': pws_title,
                            'pws_subjects': pws_subjects,
                            'student': student
                        }

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error student: ' + str(e))
# - end of ImportStudent


def ImportStudentsubject(ws_name, row_data, logfile, mapped, examyear_instance, school_instance, request):  #PR2021-05-20

    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug('-------------------  ImportStudentsubject -----------------')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('ws_name: ' + str(ws_name))
        logger.debug('examyear_instance: ' + str(examyear_instance) + ' ' + str(type(examyear_instance)))
        logger.debug('school_instance: ' + str(school_instance) + ' ' + str(type(school_instance)))


    """
    fields of awpupload13_studsubj are:
    student_id, 
    schemeitem_id,
    subject_id,
    cluster_id, 
    
    is_extra_nocount, 
    is_extra_counts, 
    has_exemption, has_sr, has_reex, has_reex03, 
    has_pok,  has_pex, 
    tv01_pescore, 
    tv01_cescore,
    tv01_segrade, tv01_srgrade, tv01_sesrgrade, 
    tv01_pegrade, tv01_cegrade, tv01_pecegrade, tv01_finalgrade, 
    tv02_cescore, tv02_cegrade,tv02_pecegrade, tv02_finalgrade, 
    tv03_cescore, tv03_cegrade, tv03_pecegrade, tv03_finalgrade, 
    tvvrst_segrade, tvvrst_cegrade, tvvrst_finalgrade,
    gradelist_sesrgrade, gradelist_pecegrade, gradelist_finalgrade
   
    """
    """
    Kandidaat_Vak.KandidaatID AS student_id, 
    Kandidaat_Vak.VakSchemaItemID AS schemeitem_id, 
    Kandidaat_Vak.VakID AS subject_id, 
    Kandidaat_Vak.ClusterID AS cluster_id, 
    IIf([Kandidaat_Vak]![IsExtraVak],1,Null) AS is_extra_nocount, 
    IIf([Kandidaat_Vak]![ExtraVakTeltMee],1,Null) AS is_extra_counts, 
    IIf([Kandidaat_Vak]![HasVrst],1,Null) AS has_exemption, 
    IIf([Kandidaat_Vak]![IsHerSe],1,Null) AS has_sr, 
    IIf([Kandidaat_Vak]![HerExamen],1,Null) AS has_reex, 
    IIf([Kandidaat_Vak]![IsHerTv03],1,Null) AS has_reex03, 
    IIf([Kandidaat_Vak]![HasBewijsKennis],1,Null) AS has_pok, 
    IIf([Kandidaat_Vak]![HasBewijsVrst],1,Null) AS has_pex, 
    Kandidaat_Vak.ScorePE AS tv01_pescore, 
    Kandidaat_Vak.ScoreCE AS tv01_cescore, 
    Kandidaat_Vak.GemidSE AS tv01_segrade, 
    Kandidaat_Vak.GemidSeHer AS tv01_srgrade, 
    Kandidaat_Vak.GemidSeFinal AS tv01_sesrgrade, 
    Kandidaat_Vak.GemidPE AS tv01_pegrade, 
    Kandidaat_Vak.GemidCSE AS tv01_cegrade, 
    Kandidaat_Vak.GemidPeCe AS tv01_pecegrade, 
    Kandidaat_Vak.Resultaat AS tv01_finalgrade, 
    Kandidaat_Vak.ScoreHer AS tv02_cescore, 
    Kandidaat_Vak.GemidHerex AS tv02_cegrade, 
    Kandidaat_Vak.GemidPeCeHer AS tv02_pecegrade, 
    Kandidaat_Vak.ResultaatHerex AS tv02_finalgrade,
    Kandidaat_Vak.ScoreTv03 AS tv03_cescore, 
    Kandidaat_Vak.GemidHerTv03 AS tv03_cegrade, 
    Kandidaat_Vak.GemidPeCeTv03 AS tv03_pecegrade, 
    Kandidaat_Vak.ResultaatTv03 AS tv03_finalgrade, 
    Kandidaat_Vak.GemidSeVrst AS tvvrst_segrade, 
    Kandidaat_Vak.GemidCseVrst AS tvvrst_cegrade, 
    Kandidaat_Vak.ResultaatVrst AS tvvrst_finalgrade, 
    [Kandidaat_Vak]![CLcijferSE] AS gradelist_sesrgrade,
    [Kandidaat_Vak]![CLcijferCE] AS gradelist_pecegrade,
    [Kandidaat_Vak]![CLcijferEIND] AS gradelist_finalgrade
    
    """
    if ws_name == 'studsubj' and row_data:
        studentsubject = None
        try:
            student, pws_title, pws_subjects = get_student_from_mapped(row_data, school_instance, mapped)
            if logging_on:
                logger.debug('student: ' + str(student))
                logger.debug('pws_title: ' + str(pws_title) + ' ' + str(len(pws_title)))
                logger.debug('pws_subjects: ' + str(pws_subjects) + ' ' + str(len(pws_subjects)))

            schemeitem = get_schemeitem_from_mapped(row_data, examyear_instance, mapped)
            if logging_on:
                logger.debug('schemeitem: ' + str(schemeitem))

            if student and schemeitem:
    # - check if studentsubject already exists in this examyear. If not: create
                studentsubject = stud_mod.Studentsubject.objects.filter(
                    student=student,
                    schemeitem=schemeitem
                ).order_by('pk').first()

                if logging_on:
                    logger.debug('studentsubject: ' + str(studentsubject))

    # - if not: create new studentsubject record
                if studentsubject is None:
                    studentsubject = stud_mod.Studentsubject(
                        student=student,
                        schemeitem=schemeitem
                    )
                    studentsubject.save(request=request)
                    logfile.append(ws_name + ' created: ' + str(studentsubject))

                    if logging_on:
                        logger.debug (ws_name + ' created = ' + str(studentsubject))

                if studentsubject:

    # - update info in studentsubject, both in new and existing records
                    # dont import cluster
                    #studentsubject.cluster = get_cluster_from_mapped(row_data, mapped)
                    studentsubject.is_extra_nocount = True if row_data.get('is_extra_nocount') == 1 else False
                    studentsubject.is_extra_counts = True if row_data.get('is_extra_counts') == 1 else False

                    # only import pws title and subjects when subjecttype = werkstuk
                    if pws_title or pws_subjects:
                        if schemeitem.subjecttype.has_pws:
                            studentsubject.pws_title = pws_title
                            studentsubject.pws_subjects = pws_subjects

                    studentsubject.has_exemption = True if row_data.get('has_exemption') == 1 else False
                    studentsubject.has_sr = True if row_data.get('has_sr') == 1 else False
                    studentsubject.has_reex = True if row_data.get('has_reex') == 1 else False
                    studentsubject.has_reex03 = True if row_data.get('has_reex03') == 1 else False
                    if row_data.get('has_pok') == 1 or row_data.get('has_pex') == 1:
                        years_valid = 10 if student.iseveningstudent or student.islexstudent else 1
                        studentsubject.pok_validthru = examyear_instance.code + years_valid  # ImportStudentsubject

                    # TODO check if these max fields are used
                    studentsubject.ex_max_segrade = row_data.get('ex_max_segrade')
                    studentsubject.ex_max_pecegrade = row_data.get('ex_max_pecegrade')
                    studentsubject.ex_max_finalgrade = row_data.get('ex_max_finalgrade')

                    studentsubject.reex_max_segrade = row_data.get('reex_max_segrade')
                    studentsubject.reex_max_pecegrade = row_data.get('reex_max_pecegrade')
                    studentsubject.reex_max_finalgrade = row_data.get('reex_max_finalgrade')

                    studentsubject.reex3_max_segrade = row_data.get('reex3_max_segrade')
                    studentsubject.reex3_max_pecegrade = row_data.get('reex3_max_pecegrade')
                    studentsubject.reex3_max_finalgrade = row_data.get('reex3_max_finalgrade')

                    studentsubject.gradelist_segrade = row_data.get('gradelist_sesrgrade')
                    studentsubject.gradelist_pecegrade = row_data.get('gradelist_pecegrade')
                    studentsubject.gradelist_finalgrade = row_data.get('gradelist_finalgrade')
                    studentsubject.save(request=request)

        except Exception as e:
            logger.error(getattr(e, 'message', str(e)))
            logfile.append('Error studentsubject: ' + str(e))

        else:
            if studentsubject:
                try:
        # - check if studentsubject has grade record of tv01. If not: create ( grade tv01 always existst
                    grade_tv01 = stud_mod.Grade.objects.filter(
                        studentsubject=studentsubject,
                        examperiod=c.EXAMPERIOD_FIRST
                    ).order_by('pk').first()
        # - create new grade_tv01 record
                    if grade_tv01 is None:
                        grade_tv01 = stud_mod.Grade(
                            studentsubject=studentsubject,
                            examperiod=c.EXAMPERIOD_FIRST
                        )
                        if logging_on:
                            logger.debug(ws_name + ' grade_tv01 created = ' + str(grade_tv01))
        # - update info in grade, both in new and existing records
                    if grade_tv01:
                        grade_tv01.pescore = get_score_from_awpimport(row_data.get('tv01_pescore'))
                        grade_tv01.cescore = get_score_from_awpimport(row_data.get('tv01_cescore'))

                        # make ovg lower case, only in se and sr
                        grade_tv01.segrade = get_grade_from_awpimport(row_data.get('tv01_segrade'))
                        grade_tv01.srgrade = get_grade_from_awpimport(row_data.get('tv01_srgrade'))
                        grade_tv01.sesrgrade = get_grade_from_awpimport(row_data.get('tv01_sesrgrade'))

                        grade_tv01.pegrade = get_grade_from_awpimport(row_data.get('tv01_pegrade'))
                        grade_tv01.cegrade = get_grade_from_awpimport(row_data.get('tv01_cegrade'))
                        grade_tv01.pecegrade = get_grade_from_awpimport(row_data.get('tv01_pecegrade'))

                        grade_tv01.finalgrade = get_grade_from_awpimport(row_data.get('tv01_finalgrade'))

                        grade_tv01.save(request=request)

                except Exception as e:
                    logger.error(getattr(e, 'message', str(e)))
                    logfile.append('Error studentsubject: ' + str(e))

                try:
    # - check if studentsubject has grade record of tv02. Create if not exists and has_reex, delete if not has_reex and exists
                    grade_tv02 = stud_mod.Grade.objects.filter(
                        studentsubject=studentsubject,
                        examperiod=c.EXAMPERIOD_SECOND
                    ).order_by('pk').first()
                    if studentsubject.has_reex:
                        if grade_tv02 is None:
                            grade_tv02 = stud_mod.Grade(
                                studentsubject=studentsubject,
                                examperiod=c.EXAMPERIOD_SECOND
                            )
                        if grade_tv02:
                            grade_tv02.cescore = get_score_from_awpimport(row_data.get('tv02_cescore'))

                            grade_tv02.cegrade = get_grade_from_awpimport(row_data.get('tv02_cegrade'))
                            grade_tv02.pecegrade = get_grade_from_awpimport(row_data.get('tv02_pecegrade'))
                            grade_tv02.finalgrade = get_grade_from_awpimport(row_data.get('tv02_finalgrade'))

                            grade_tv02.save(request=request)
                    else:
                        if grade_tv02:
                            grade_tv02.delete(request=request)

                except Exception as e:
                    logger.error(getattr(e, 'message', str(e)))
                    logfile.append('Error studentsubject: ' + str(e))

                try:
    # - check if studentsubject has grade record of tv03. Create if not exists and has_reex, delete if not has_reex and exists
                    grade_tv03 = stud_mod.Grade.objects.filter(
                        studentsubject=studentsubject,
                        examperiod=c.EXAMPERIOD_THIRD
                    ).order_by('pk').first()
                    if studentsubject.has_reex03:
                        if grade_tv03 is None:
                            grade_tv03 = stud_mod.Grade(
                                studentsubject=studentsubject,
                                examperiod=c.EXAMPERIOD_THIRD
                            )
                        if grade_tv03:
                            grade_tv03.cescore = get_score_from_awpimport(row_data.get('tv03_cescore'))

                            grade_tv03.cegrade = get_grade_from_awpimport(row_data.get('tv03_cegrade'))
                            grade_tv03.pecegrade = get_grade_from_awpimport(row_data.get('tv03_pecegrade'))
                            grade_tv03.finalgrade = get_grade_from_awpimport(row_data.get('tv03_finalgrade'))

                            grade_tv03.save(request=request)
                    else:
                        if grade_tv03:
                            grade_tv03.delete(request=request)

                except Exception as e:
                    logger.error(getattr(e, 'message', str(e)))
                    logfile.append('Error studentsubject: ' + str(e))
                try:
# - check if studentsubject has exemoption record . Create if not exists and has_exemption, delete if not has_exemption and exists
                    grade_tvexem = stud_mod.Grade.objects.filter(
                        studentsubject=studentsubject,
                        examperiod=c.EXAMPERIOD_EXEMPTION
                    ).order_by('pk').first()
                    if studentsubject.has_exemption:
                        if grade_tvexem is None:
                            grade_tvexem = stud_mod.Grade(
                                studentsubject=studentsubject,
                                examperiod=c.EXAMPERIOD_EXEMPTION
                            )
                        if grade_tvexem:
                            grade_tvexem.segrade = row_data.get('tvvrst_segrade')
                            grade_tvexem.pecegrade = row_data.get('tvvrst_cegrade')
                            grade_tvexem.finalgrade = row_data.get('tvvrst_finalgrade')
                            grade_tvexem.save(request=request)
                    else:
                        if grade_tvexem:
                            grade_tvexem.delete(request=request)

                except Exception as e:
                    logger.error(getattr(e, 'message', str(e)))
                    logfile.append('Error studentsubject: ' + str(e))

# - end of ImportStudentsubject


def get_score_from_awpimport(value): # PR2021-09-20
    score_int = None
    try:
        if value is not None:
            if isinstance(value, int):
                score_int = int(value)
            elif isinstance(value, str) and str.isdigit(value):
                score_int = int(value)
    except Exception as e:
        logger.error(getattr(e, 'message', str(e)))
    return score_int


def get_grade_from_awpimport(value): # PR2021-09-20

    # PR2021-11-22 from now grades are saved with dots instead of comma's,
    #  so they can be used by Decimal() without having to convert to dots
    #  replace by comma's when printing gradelist and reports

    grade_str = None
    try:
        if value is not None:
            value_with_dot = value.replace(',', '.')
            grade_str = str(value_with_dot).lower()
    except Exception as e:
        logger.error(getattr(e, 'message', str(e)))
    return grade_str
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@

def get_department_from_mapped(row_data, examyear, mapped):  # PR2021-05-04
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_department_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear: ' + str(examyear))
        logger.debug('mapped: ' + str(mapped.get('department', '')))

    # mapped[department]: {1: 1, 2: 2, 3: 3}
    # key = AWP afdelingAID, value = awpr department.pk
    department = None
    if row_data and mapped:
        mapped_departments = mapped.get('department')
        if mapped_departments:
            awp_afdeling_aid = row_data.get('department_id')
            department_pk = mapped_departments.get(awp_afdeling_aid)
            if logging_on:
                logger.debug('department_pk: ' + str(department_pk))

            if department_pk:
                department = sch_mod.Department.objects.get_or_none(
                    pk=department_pk,
                    examyear=examyear
                )
    if logging_on:
        logger.debug('department: ' + str(department))
        logger.debug(' ----- end of get_department_from_mapped -----')

    return department
# - end of get_department_from_mapped


def get_level_from_mapped(row_data, examyear, department, mapped):  # PR2021-05-04
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_level_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear: ' + str(examyear))
        logger.debug('mapped: ' + str(mapped.get('depbase', '')))
    # mapped[level]: {1: 3, 2: 2, 3: 1}
    # key = AWP StudierichtingAID, value = awpr level.pk

    level = None
    if department and mapped:
        mapped_levels = mapped.get('level')

        # skip Havo/Vwo: Havo / Vwo: don't have level
        if department.base.code.lower() == 'vsbo':
            # Vsbo: level is stored in AWP table Studierichting
            awp_aid = row_data.get('level_id')
            if awp_aid:
                level_pk = mapped_levels.get(awp_aid)
                level = subj_mod.Level.objects.get_or_none(
                    pk=level_pk,
                    examyear=examyear
                )
                if logging_on:
                    logger.debug('awp_aid: ' + str(awp_aid))
                    logger.debug('level_pk: ' + str(level_pk))

    if logging_on:
        logger.debug('level: ' + str(level))
        logger.debug(' ----- end of get_level_from_mapped -----')

    return level
# - end of get_level_from_mapped


def get_sector_from_mapped(row_data, examyear, department, mapped):  # PR2021-05-04
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_sector_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear: ' + str(examyear))
        logger.debug('mapped: ' + str(mapped.get('sector', '')))

    # row_data: {'scheme_id': 1, 'department_id': 1, 'sector_id': 1, 'level_id': 1}

    #  mapped[sector]: {'sdr_5': 7, 'sdr_6': 6, 'sdr_7': 5, 'sdr_8': 4, 'sct_1': 2, 'sct_2': 1, 'sct_3': 3}
    # Note when Vsbo: sector has prefix 'sct' (stored in awp table 'Sector'
    # when Havo / Vwo: sector is stored in AWP table Studierichting and has prefix 'sdr

    sector = None
    if department and mapped:
        mapped_sectors = mapped.get('sector')

        if department.base.code.lower() == 'vsbo':
            # Vsbo: sector is stored in AWP table 'Sector'
            awp_aid = row_data.get('sector_id')
            prefix = 'sct_'
        else:
            # Havo / Vwo:: sector is stored in AWP table 'Studierichting'
            awp_aid = row_data.get('level_id')
            prefix = 'sdr_'

        if awp_aid:
            awp_aid_with_prefix = prefix + str(awp_aid)
            sector_pk = mapped_sectors.get(awp_aid_with_prefix)
            sector = subj_mod.Sector.objects.get_or_none(
                    pk=sector_pk,
                    examyear=examyear
            )
            if logging_on:
                logger.debug('sector_aid_with_prefix: ' + str(awp_aid_with_prefix))
                logger.debug('sector_pk: ' + str(sector_pk))

    if logging_on:
        logger.debug('sector: ' + str(sector))
        logger.debug(' ----- end of get_sector_from_mapped -----')

    return sector
# - end of get_sector_from_mapped


def get_scheme_from_mapped(row_data, examyear_instance, mapped):  # PR2021-05-04
    logging_on = False # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_scheme_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear_instance: ' + str(examyear_instance) + ' ' + str(type(examyear_instance)))

    # mapped[scheme]: {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12, 13: 13, 14: 14, 15: 15, 16: 16, 17: 17}
    # key = AWP VakschemaAID, value = awpr scheme.pk
    # row_data: {'schemeitem_id': 4, 'scheme_id': 1, 'subject_id': 5, 'subjecttype_id': 1, 'gradetype_id': 1, 'weight_se': 1, 'is_mandatory': 1, 'is_combi': 1}
    scheme = None
    if row_data and mapped:
        mapped_schemes = mapped.get('scheme')
        awp_vks_aid = row_data.get('scheme_id')
        if logging_on:
            logger.debug('mapped_schemes: ' + str(mapped_schemes))
            logger.debug('awp_vks_aid: ' + str(awp_vks_aid))

        if mapped_schemes and awp_vks_aid:
            scheme_pk = mapped_schemes.get(awp_vks_aid)
            if logging_on:
                logger.debug('scheme_pk: ' + str(scheme_pk))

            if scheme_pk:
                scheme = subj_mod.Scheme.objects.get_or_none(
                    pk=scheme_pk,
                    department__examyear=examyear_instance
            )
    if logging_on:
        logger.debug('scheme: ' + str(scheme))
        logger.debug(' ----- end of get_scheme_from_mapped -----')

    return scheme
# - end of get_scheme_from_mapped


def get_subject_from_mapped(row_data, examyear_instance, mapped):  # PR2021-05-04
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_subject_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear_instance: ' + str(examyear_instance))

    # mapped[subject]: {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12, 13: 13, 14: 14, 15: 15, 16: 16, 17: 17}
    # key = AWP VakschemaAID, value = awpr scheme.pk
    # row_data: {'schemeitem_id': 4, 'scheme_id': 1, 'subject_id': 5, 'subjecttype_id': 1, 'gradetype_id': 1, 'weight_se': 1, 'is_mandatory': 1, 'is_combi': 1}

    subject = None
    if row_data and mapped:
        mapped_subjects = mapped.get('subject')
        awp_vak_aid = row_data.get('subject_id')
        if logging_on:
            logger.debug('mapped_subjects: ' + str(mapped_subjects))
            logger.debug('awp_vks_aid: ' + str(awp_vak_aid))

        if mapped_subjects and awp_vak_aid:
            subject_pk = mapped_subjects.get(awp_vak_aid)
            if logging_on:
                logger.debug('subject_pk: ' + str(subject_pk))

            if subject_pk:
                subject = subj_mod.Subject.objects.get_or_none(
                    pk=subject_pk,
                    examyear=examyear_instance
            )
    if logging_on:
        logger.debug('subject: ' + str(subject))
        logger.debug(' ----- end of get_subject_from_mapped -----')

    return subject
# - end of get_subject_from_mapped


def get_subjecttype_from_mapped(row_data, scheme, mapped):  # PR2021-05-04
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_subjecttype_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))

    subjecttype = None
    if row_data and mapped:
        mapped_sjtpbase = mapped.get('sjtpbase')
        awp_vaktype_aid = row_data.get('subjecttype_id')

        if logging_on:
            logger.debug('mapped_subjecttypes: ' + str(mapped_sjtpbase))
            logger.debug('awp_vaktype_aid: ' + str(awp_vaktype_aid) + ' ' + str(type(awp_vaktype_aid)))

        if mapped_sjtpbase and awp_vaktype_aid:
            sjtpbase_pk = mapped_sjtpbase.get(awp_vaktype_aid)
            if logging_on:
                logger.debug('sjtpbase_pk: ' + str(sjtpbase_pk))

            if sjtpbase_pk:
                subjecttypebase = subj_mod.Subjecttypebase.objects.get_or_none(
                    pk=sjtpbase_pk)
                if logging_on:
                    logger.debug('subjecttypebase: ' + str(subjecttypebase))

                if subjecttypebase:
                    subjecttype = subj_mod.Subjecttype.objects.filter(
                        base=subjecttypebase,
                        scheme=scheme
                    ).order_by('pk').first()

    if logging_on:
        logger.debug('subjecttype: ' + str(subjecttype))
        logger.debug(' ----- end of get_subjecttype_from_mapped -----')

    return subjecttype
# - end of get_subjecttype_from_mapped


def get_package_from_mapped(row_data, mapped):  # PR2021-05-05
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_package_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))

    # mapped[package]: {1: 1, 2: 1, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2, 9: 3, 10: 3, ...
    # row_data: {'package_id': 1, 'subject_id': 2, 'schemeitem_id': 134}
    package = None
    if row_data and mapped:
        mapped_packages = mapped.get('package')
        awp_vks_aid = row_data.get('package_id')
        if logging_on:
            logger.debug('awp_vks_aid: ' + str(awp_vks_aid))

        if mapped_packages and awp_vks_aid:
            package_pk = mapped_packages.get(awp_vks_aid)
            if logging_on:
                logger.debug('package_pk: ' + str(package_pk))

            if package_pk:
                package = subj_mod.Package.objects.get_or_none(
                    pk=package_pk
                )
    if logging_on:
        logger.debug('package: ' + str(package))
        logger.debug(' ----- end of get_package_from_mapped -----')

    return package
# - end of get_package_from_mapped


def get_schemeitem_from_mapped(row_data, examyear_instance, mapped):  # PR2021-05-05
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_schemeitem_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('examyear_instance: ' + str(examyear_instance))

    # mapped[schemeitem]: {1: 1, 2: 1, 3: 2, 4: 2, 5: 2, 6: 2, 7: 2, 8: 2, 9: 3, 10: 3, ...
    # row_data: {'schemeitem_id': 4, 'scheme_id': 1, 'subject_id': 5, 'subjecttype_id': 1, 'gradetype_id': 1, 'weight_se': 1, 'is_mandatory': 1, 'is_combi': 1}
    schemeitem = None
    if row_data and mapped:
        mapped_schemeitems = mapped.get('schemeitem')
        awp_vsi_aid = row_data.get('schemeitem_id')
        if logging_on:
            logger.debug('mapped_schemeitems: ' + str(mapped_schemeitems))
            logger.debug('awp_vsi_aid: ' + str(awp_vsi_aid))

        if mapped_schemeitems and awp_vsi_aid:
            schemeitem_pk = mapped_schemeitems.get(awp_vsi_aid)
            if logging_on:
                logger.debug('schemeitem_pk: ' + str(schemeitem_pk))

            if schemeitem_pk:
                schemeitem = subj_mod.Schemeitem.objects.get_or_none(
                    pk=schemeitem_pk,
                    scheme__department__examyear=examyear_instance
            )
    if logging_on:
        logger.debug('schemeitem: ' + str(schemeitem))
        logger.debug(' ----- end of get_schemeitem_from_mapped -----')

    return schemeitem
# - end of get_schemeitem_from_mapped



def get_cluster_from_mapped(row_data, mapped):  # PR2021-05-20
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_cluster_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))

    #  mapped[student = { KandidaatID: = { 'pk': student.pk, 'pws_title': pws_title, 'pws_subjects': pws_subjects }

    cluster = None
    if row_data and mapped:
        mapped_clusters = mapped.get('cluster')

        awp_cluster_id = row_data.get('cluster_id')
        if logging_on:
            logger.debug('mapped_clusters: ' + str(mapped_clusters))
            logger.debug('awp_cluster_id: ' + str(awp_cluster_id))

        if mapped_clusters and awp_cluster_id:
            cluster = mapped_clusters.get(awp_cluster_id)

    if logging_on:
        logger.debug('cluster: ' + str(cluster))
        logger.debug(' ----- end of get_cluster_from_mapped -----')

    return cluster


# - end of get_student_from_mapped
def get_student_from_mapped(row_data, school_instance, mapped):  # PR2021-05-20
    logging_on = False  # s.LOGGING_ON
    if logging_on:
        logger.debug(' ----- get_student_from_mapped -----')
        logger.debug('row_data: ' + str(row_data))
        logger.debug('school_instance: ' + str(school_instance))

    #  mapped[student = { KandidaatID: = { 'pk': student.pk, 'pws_title': pws_title, 'pws_subjects': pws_subjects }

    student, pws_title, pws_subjects = None, None, None
    if row_data and mapped:
        mapped_students = mapped.get('student')

        awp_kandidaat_aid = row_data.get('student_id')
        if logging_on:
            logger.debug('mapped_students: ' + str(mapped_students))
            logger.debug('awp_kandidaat_aid: ' + str(awp_kandidaat_aid))

        if mapped_students and awp_kandidaat_aid:
            mapped_student = mapped_students.get(awp_kandidaat_aid)
            if mapped_student:
                if logging_on:
                    logger.debug('mapped_student: ' + str(mapped_student))
                student = mapped_student.get('student')
                pws_title = mapped_student.get('pws_title')
                pws_subjects = mapped_student.get('pws_subjects')

    if logging_on:
        logger.debug('student: ' + str(student))
        logger.debug('student: ' + str(student))
        logger.debug(' ----- end of get_student_from_mapped -----')

    return student, pws_title, pws_subjects
# - end of get_student_from_mapped

# ===================== Functions

def get_depbase_id_list_from_nameslist(names_list, mapped):  # PR2018-12-12  PR2020-12-13
    # PR2018-12-12 function replaces 'tkl,pkl,pbl' with ';1;2;3;'
    # PR2020-12-09 changed to arrayfield 'depbases' = [1,2,3]
    # mapped_sectorlist_dict is a separate function that gets the baseID's from this examyear
    # mapped_baselist = {'vsbo': 4, 'havo': 5,'vwo': 6}
    mapped_baselist = mapped.get('depbase')
    base_id_list = []
    if names_list:
        array = names_list.split(';')
        if array:
            for abbrev in array:
                # look up abbrev in mapped_baselist
                if abbrev:
                    base_id = mapped_baselist.get(abbrev.lower())
                    if base_id:
                        base_id_list.append(base_id)
    return base_id_list


def create_scheme_name(depbase_code, level_abbrev, sector_abbrev ):
    # PR2018-11-09 create scheme-name i.e.: 'vsbo - tkl - tech' PR2020-12-13
    scheme_name = ''
    if depbase_code:
        scheme_name = depbase_code
    if level_abbrev and level_abbrev.lower() != 'none':
        scheme_name = scheme_name + " - " + level_abbrev
    if sector_abbrev:
        scheme_name = scheme_name + " - " + sector_abbrev
    return scheme_name


def get_examyear(country, examyear_int):
    # lookup examyear of country PR2020-12-14
    # examyear is integer field of table Examyear
    examyear = None
    if country and examyear_int:
        examyear = sch_mod.Examyear.objects.filter(
            country=country,
            code=examyear_int
        ).order_by('pk').first()
    return examyear


def get_subjecttype(subjtype_name, examyear):
    # check if subjecttype already exists
    subjecttype = None
    if subjtype_name and examyear:
        subjecttype = subj_mod.Subjecttype.objects.filter(
            examyear=examyear,
            name__iexact=subjtype_name
        ).order_by('pk').first()
    return subjecttype


def get_subject(subject_code, examyear):
    # check if subject already exists
    subject = None
    if subject_code and examyear:
        subject =subj_mod.Subject.objects.filter(
            examyear=examyear,
            base__code__iexact=subject_code
        ).order_by('pk').first()
    return subject


def get_scheme(depbase_code, level_abbrev, sector_abbrev, examyear):
    # check if scheme already exists
    #logger.debug('------- get_scheme -------------')

    scheme = None
    if depbase_code and examyear:
        crit = Q(department__examyear=examyear) & \
               Q(department__base__code__iexact=depbase_code)
        if level_abbrev and not level_abbrev.lower() == 'none':
            crit.add(Q(level__abbrev__iexact=level_abbrev), crit.connector)
        if sector_abbrev and not sector_abbrev.lower() == 'none':
            crit.add(Q(sector__abbrev__iexact=sector_abbrev), crit.connector)
        scheme = subj_mod.Scheme.objects.filter(crit).order_by('pk').first()
    return scheme


def get_schemeitem(scheme, subject, subjecttype):
    # get existing schemeitem PR2020-12-13
    schemeitem = None
    if scheme and subject and subjecttype:
        schemeitem = subj_mod.Schemeitem.objects.filter(
            scheme=scheme,
            subject=subject,
            subjecttype=subjecttype
        ).order_by('pk').first()
    return schemeitem


def get_package(school, scheme, package_name):
    # get existing schemeitem PR2020-12-13
    package = None
    if school and scheme and package_name:
        package = subj_mod.Package.objects.filter(
            school=school,
            scheme=scheme,
            name__iexact=package_name
        ).order_by('pk').first()
    return package


def get_packageitem(package, schemeitem):
    # get existing packageitem PR2020-12-13
    packageitem = None
    if package and schemeitem:
        packageitem = subj_mod.Packageitem.objects.filter(
            package=package,
            schemeitem=schemeitem
        ).order_by('pk').first()
    return packageitem



def get_birthcountry(name):
    # check if subject already exists
    birthcountry = None
    if name:
        birthcountry = stud_mod.Birthcountry.objects.filter(
            name__iexact=name
        ).order_by('pk').first()
    return birthcountry


def get_birthplace(birthcountry, birthplace_name):
    # check if subject already exists
    birthplace = None
    if birthcountry and birthplace_name:
        birthplace = stud_mod.Birthplace.objects.filter(
            birthcountry=birthcountry,
            name__iexact=birthplace_name
        ).order_by('pk').first()
    return birthplace


def get_schoolbase(country, code):
    # get existing schoolbase PR2020-12-14
    schoolbase = None
    if country and code:
        schoolbase = sch_mod.Schoolbase.objects.filter(
            country=country,
            code__iexact=code
        ).order_by('pk').first()
    return schoolbase


def get_school(schoolbase, examyear):
    # get existing school from request_user PR2020-12-13
    school = None
    if schoolbase and examyear:
        school = sch_mod.School.objects.filter(
            base=schoolbase,
            examyear=examyear
        ).order_by('pk').first()
    return school
